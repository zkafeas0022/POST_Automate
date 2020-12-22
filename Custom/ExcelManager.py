# !/usr/bin/python
# -*-coding: utf-8-*-
import natsort
import openpyxl
from operator import itemgetter
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Font, colors

import sys
reload(sys)
sys.setdefaultencoding('UTF-8')


class ExcelManager:

    def __init__(self):
        self.wb = None
        self.sheet_names = None
        self.fileName = None

    def open_excel_file(self, filename):
        self.wb = openpyxl.load_workbook(filename)
        self.fileName = filename
        self.sheet_names = self.wb.sheetnames
        return self.wb

    def get_sheets_name(self):
        return self.sheet_names

    def get_row_count(self, sheet_name):
        sheet = self.wb.get_sheet_by_name(sheet_name)
        return sheet.max_row

    def get_column_count(self, sheet_name):
        sheet = self.wb.get_sheet_by_name(sheet_name)
        return sheet.max_column

    def find_row_by_value(self, sheet_name, value):
        row_result = None
        sheet = self.wb.get_sheet_by_name(sheet_name)
        max_row = sheet.max_row
        max_column = sheet.max_column
        for row_index in range(1, max_row + 1):
            for col_index in range(1, max_column + 1):
                cell = get_column_letter(col_index) + str(row_index)
                cell_value = sheet[cell].value
                if str(cell_value) == str(value):
                    row_result = row_index
                    break
        return row_result

    def find_column_by_value(self, sheet_name, value):
        column_result = None
        sheet = self.wb.get_sheet_by_name(sheet_name)
        max_row = sheet.max_row
        max_column = sheet.max_column
        for row_index in range(1, max_row + 1):
            for col_index in range(1, max_column + 1):
                cell = get_column_letter(col_index) + str(row_index)
                cell_value = sheet[cell].value
                if str(cell_value) == str(value):
                    column_result = col_index
                    break
        return column_result

    def get_cell_value_by_position(self, sheet_name, row, column, encoding="UTF-8"):
        sheet = self.wb.get_sheet_by_name(sheet_name)
        try:
            cell_value = sheet.cell(row=int(row), column=int(column)).value
            if cell_value is not None:
                value = cell_value.encode(encoding, 'ignore').decode(encoding)
            else:
                value = cell_value
            print value
        except Exception, e:
            raise str(e)
        return value

    def get_cell_value_by_row(self, sheet_name, row):
        sheet = self.wb.get_sheet_by_name(sheet_name)
        max_column = sheet.max_column
        data = []
        for col_index in range(1, max_column):
            cell_value = self.get_cell_value_by_position(sheet_name=sheet_name, row=row, column=col_index)
            data.append(str(cell_value))
        return data

    def get_cell_value_by_column(self, sheet_name, column):
        sheet = self.wb.get_sheet_by_name(str(sheet_name))
        max_row = sheet.max_row
        data = []
        for row_index in range(2, (max_row + 1)):
            cell_value = self.get_cell_value_by_position(sheet_name=sheet_name, row=row_index, column=column)
            data.append(str(cell_value))
        return data

    def read_excel_file(self, sheet_name, include_empty_cells=True):
        sheet = self.wb.get_sheet_by_name(sheet_name)
        max_row = sheet.max_row
        max_column = sheet.max_column
        data = {}
        for row_index in range(1, max_row):
            for col_index in range(1, max_column):
                cell = get_column_letter(col_index) + str(row_index)
                value = sheet[cell].value
                data[cell] = str(value)
        if include_empty_cells is True:
            sorted_data = natsort.natsorted(data.items(), key=itemgetter(0))
            return sorted_data
        else:
            data = dict([(k, v) for (k, v) in data.items() if v])
            ordered_data = natsort.natsorted(data.items(), key=itemgetter(0))
            return ordered_data

    def get_next_cell_value_by_condition(self, sheet_name, value):
        sheet = self.wb.get_sheet_by_name(sheet_name)
        max_row = sheet.max_row
        max_column = sheet.max_column
        for row in range(1, max_row + 1):
            for column in range(1, max_column):
                value_column = get_column_letter(column)
                cell_value = sheet[value_column + str(row)].value
                if value == cell_value:
                    value_column = get_column_letter(column + 1)
                    return str(sheet[value_column + str(row)].value)

    def write_excel_value(self, sheet_name, row, col, val, font_color, bg_color):
        sheet = self.wb.get_sheet_by_name(sheet_name)
        try:
            if str.upper(font_color) == "RED":
                font_style = Font(color="E60000")
            elif str.upper(font_color) == "BLUE":
                font_style = Font(color="0000CC")
            elif str.upper(font_color) == "GREEN":
                font_style = Font(color="006600")
            elif str.upper(font_color) == "BLACK":
                font_style = Font(color="0D0D0D")
            else:
                font_style = Font(color=str.upper(font_color))

            if str.upper(bg_color) != "" and str.upper(bg_color) != "NO FILL":
                fill_color = PatternFill("solid", fgColor=str.upper(bg_color))
            elif str.upper(bg_color) == "NO FILL":
                fill_color = "NO FILL"
            else:
                fill_color = PatternFill("solid", fgColor=colors.YELLOW)

            cell = sheet.cell(row=int(row), column=int(col))
            cell.font = font_style
            cell.value = val.decode('cp874')

            if fill_color != "NO FILL":
                cell.fill = fill_color

            self.wb.save(self.fileName)

            return True
        except ValueError as e:
            print "can not write value on cell at row " + str(row) + " column " + str(col) + "," + str(e)
            return False

    @staticmethod
    def create_excel_file(file_name_with_path, sheet_name):
        if file_name_with_path != "" and sheet_name != "":
            try:
                book = openpyxl.Workbook()
                arr_original_sheet = book.get_sheet_names()
                for i_sheet in range(0, len(arr_original_sheet)):
                    del_sheet = book.get_sheet_by_name(arr_original_sheet[i_sheet])
                    book.remove_sheet(del_sheet)

                arr_new_sheet = str.split(sheet_name, ",")
                for i_sheet in range(0, len(arr_new_sheet)):
                    book.create_sheet(arr_new_sheet[i_sheet], i_sheet)

                book.save(file_name_with_path)
                return True
            except IOError:
                return False
